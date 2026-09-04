from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.accounts.models import User
from apps.catalog.models import Product
from apps.core.models import AppSettings
from apps.inventory.models import StockMovement, Warehouse
from apps.inventory.services import move_stock
from apps.sales import services as sales_services


def make_warehouses():
    main = Warehouse.objects.create(name="Основной склад", code="MAIN", kind=Warehouse.Kind.MAIN, is_sellable=False)
    shop = Warehouse.objects.create(name="Магазин", code="SHOP", kind=Warehouse.Kind.SHOP, is_sellable=True)
    return main, shop


class ExcelExportTests(TestCase):
    def setUp(self):
        AppSettings.load()
        self.main, self.shop = make_warehouses()
        self.owner = User.objects.create_user(username="owner3", password="x", role=User.Role.OWNER)
        self.seller = User.objects.create_user(username="seller3", password="x", role=User.Role.SELLER)
        self.product = Product.objects.create(
            name="Фильтр экспортный", sku="EXP-1", sale_price=5000, purchase_price=2000, avg_cost=2000,
        )
        move_stock(product=self.product, warehouse=self.shop, qty_delta=Decimal("10"),
                   reason=StockMovement.Reason.RECEIPT, unit_cost=Decimal("2000"))
        sales_services.create_sale(
            warehouse=self.shop, seller=self.seller,
            items_data=[{"product": self.product, "quantity": Decimal("2"), "final_price": Decimal("4000")}],
            payments_data=[{"method": "cash", "amount": Decimal("8000")}],
        )

    def test_daily_report_export_is_valid_xlsx_with_expected_data(self):
        client = APIClient()
        client.force_authenticate(user=self.owner)
        resp = client.get("/api/v1/reports/daily/export/")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheetml", resp["Content-Type"])
        self.assertIn("attachment", resp["Content-Disposition"])

        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        self.assertTrue(any("Фильтр экспортный" in c for c in cells))
        self.assertTrue(any("EXP-1" in c for c in cells))

    def test_seller_export_hides_cost_columns(self):
        client = APIClient()
        client.force_authenticate(user=self.seller)
        resp = client.get("/api/v1/reports/daily/export/")
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        self.assertFalse(any(c == "Валовая прибыль" for c in cells))

    def test_stock_export_is_valid_xlsx(self):
        client = APIClient()
        client.force_authenticate(user=self.owner)
        resp = client.get("/api/v1/reports/stock/export/")
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        self.assertTrue(any("Фильтр экспортный" in c for c in cells))

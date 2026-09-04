"""Excel (.xlsx) export builders for reports.

Kept separate from views.py/services.py so the openpyxl formatting code
doesn't clutter the data-aggregation logic.
"""
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(color="6B7280", italic=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = "#,##0 ₸"


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _num(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return value
    return value


def _header_row(ws, row, labels):
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


PAYMENT_LABELS = {"cash": "Наличные", "kaspi_qr": "Kaspi QR", "card": "Карта", "transfer": "Перевод"}


def build_daily_report_workbook(data, can_see_cost=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "Дневной отчёт"

    ws["A1"] = "AutoZap — дневной отчёт"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Дата: {data['date']}"
    ws["A2"].font = SUBTITLE_FONT
    row = 4

    finance = data["finance"]
    summary_rows = [
        ("Выручка", finance["revenue"], MONEY_FMT),
        ("Чеков", finance["sales_count"], "0"),
        ("Средний чек", finance["avg_check"], MONEY_FMT),
        ("Скидки предоставлено", finance["discount_total"], MONEY_FMT),
        ("Средняя скидка, %", finance["avg_discount_pct"], "0.0"),
        ("Возвраты", finance["returns_amount"], MONEY_FMT),
        ("Чистая выручка", finance["net_revenue"], MONEY_FMT),
    ]
    if can_see_cost and finance.get("profit") is not None:
        summary_rows += [
            ("Себестоимость", finance.get("cost_total"), MONEY_FMT),
            ("Валовая прибыль", finance.get("profit"), MONEY_FMT),
            ("Рентабельность, %", finance.get("margin_pct"), "0.0"),
        ]

    ws.cell(row=row, column=1, value="Показатель").font = BOLD
    ws.cell(row=row, column=2, value="Значение").font = BOLD
    row += 1
    for label, value, fmt in summary_rows:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=_num(value))
        cell.number_format = fmt
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="По способам оплаты").font = BOLD
    row += 1
    _header_row(ws, row, ["Способ оплаты", "Сумма"])
    row += 1
    for p in finance["revenue_by_payment"]:
        ws.cell(row=row, column=1, value=PAYMENT_LABELS.get(p["method"], p["method"]))
        c = ws.cell(row=row, column=2, value=_num(p["amount"]))
        c.number_format = MONEY_FMT
        row += 1
    if not finance["revenue_by_payment"]:
        ws.cell(row=row, column=1, value="Нет данных")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Проданные товары").font = BOLD
    row += 1
    headers = ["Товар", "Код", "Кол-во", "По прайсу", "Факт", "Скидка"]
    if can_see_cost:
        headers += ["Себестоимость", "Прибыль"]
    _header_row(ws, row, headers)
    items_header_row = row
    row += 1
    for i in data["items"]:
        ws.cell(row=row, column=1, value=i["product__name"])
        ws.cell(row=row, column=2, value=i["product__sku"])
        ws.cell(row=row, column=3, value=_num(i["qty"]))
        c4 = ws.cell(row=row, column=4, value=_num(i["amount_base"]))
        c4.number_format = MONEY_FMT
        c5 = ws.cell(row=row, column=5, value=_num(i["amount_fact"]))
        c5.number_format = MONEY_FMT
        c6 = ws.cell(row=row, column=6, value=_num(i["discount"]))
        c6.number_format = MONEY_FMT
        if can_see_cost:
            c7 = ws.cell(row=row, column=7, value=_num(i.get("cost")))
            c7.number_format = MONEY_FMT
            c8 = ws.cell(row=row, column=8, value=_num(i.get("profit")))
            c8.number_format = MONEY_FMT
        row += 1
    if not data["items"]:
        ws.cell(row=row, column=1, value="Продаж не было")
        row += 1
    items_end_row = row - 1

    if data["sellers"] and len(data["sellers"]) > 1:
        row += 1
        ws.cell(row=row, column=1, value="По продавцам").font = BOLD
        row += 1
        seller_headers = ["Продавец", "Чеков", "Выручка", "Скидки"]
        if can_see_cost:
            seller_headers.append("Прибыль")
        _header_row(ws, row, seller_headers)
        row += 1
        for s in data["sellers"]:
            name = f"{s.get('seller__first_name', '')} {s.get('seller__last_name', '')}".strip() or "—"
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=s["count"])
            c3 = ws.cell(row=row, column=3, value=_num(s["revenue"]))
            c3.number_format = MONEY_FMT
            c4 = ws.cell(row=row, column=4, value=_num(s["discount"]))
            c4.number_format = MONEY_FMT
            if can_see_cost:
                c5 = ws.cell(row=row, column=5, value=_num(s.get("profit")))
                c5.number_format = MONEY_FMT
            row += 1

    if items_end_row >= items_header_row + 1:
        ws.auto_filter.ref = f"A{items_header_row}:{get_column_letter(len(headers))}{items_end_row}"

    _autosize(ws, [32, 12, 10, 14, 14, 12, 14, 14])
    ws.freeze_panes = "A5"
    return wb


def build_stock_workbook(rows):
    """rows: iterable of dicts shaped like StockSerializer output."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"

    ws["A1"] = "AutoZap — остатки на складах"
    ws["A1"].font = TITLE_FONT
    row = 3

    headers = ["Товар", "Код", "Склад", "Кол-во", "Мин. остаток", "Статус"]
    _header_row(ws, row, headers)
    header_row = row
    row += 1

    status_labels = {"out": "Нет в наличии", "low": "Дефицит", "warning": "Заканчивается", "ok": "В наличии"}
    for r in rows:
        ws.cell(row=row, column=1, value=r["product_name"])
        ws.cell(row=row, column=2, value=r["sku"])
        ws.cell(row=row, column=3, value=r["warehouse_name"])
        ws.cell(row=row, column=4, value=_num(r["quantity"]))
        ws.cell(row=row, column=5, value=r["min_stock"])
        status_cell = ws.cell(row=row, column=6, value=status_labels.get(r["status"], r["status"]))
        if r["status"] in ("out", "low"):
            status_cell.font = Font(color="B91C1C", bold=True)
        elif r["status"] == "warning":
            status_cell.font = Font(color="B45309")
        row += 1
    end_row = row - 1

    if end_row >= header_row + 1:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{end_row}"

    _autosize(ws, [34, 12, 16, 10, 12, 16])
    ws.freeze_panes = "A4"
    return wb
